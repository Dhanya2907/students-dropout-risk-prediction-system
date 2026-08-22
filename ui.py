import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os


# ============================================================
# CONFIGURATION
# ============================================================

FILE_NAME = "student_details.xlsx"

BG = "#F4F7FB"
WHITE = "#FFFFFF"

SIDEBAR = "#172554"
SIDEBAR_HOVER = "#1E3A8A"

PRIMARY = "#2563EB"
PRIMARY_DARK = "#1D4ED8"

TEXT = "#172033"
GRAY = "#64748B"
LIGHT_GRAY = "#E2E8F0"

GREEN = "#16A34A"
GREEN_LIGHT = "#F0FDF4"

ORANGE = "#EA580C"
ORANGE_LIGHT = "#FFF7ED"

RED = "#DC2626"
RED_LIGHT = "#FEF2F2"

BLUE_LIGHT = "#EFF6FF"


# ============================================================
# GLOBAL VARIABLES
# ============================================================

student_id_entry = None
student_name_entry = None
attendance_entry = None
study_hours_entry = None
internal_entry = None
assignment_entry = None
gpa_entry = None
backlog_entry = None

student_dropdown = None

content = None

prediction_result_frame = None
prediction_recommendation_frame = None


# ============================================================
# MAIN WINDOW
# ============================================================

root = tk.Tk()

root.title("Smart Student Dropout Risk Prediction System")

root.geometry("1250x760")

root.minsize(1050, 650)

root.configure(bg=BG)

# Normal Windows toolbar is enabled
# Minimize / Maximize / Close buttons will appear normally.


# ============================================================
# TERMINAL LOG
# ============================================================

def terminal_log(message):
    print("[SYSTEM]", message)


# ============================================================
# EXIT
# ============================================================

def exit_application():

    terminal_log("Exit button clicked.")

    answer = messagebox.askyesno(
        "Exit Application",
        "Are you sure you want to exit?"
    )

    if answer:
        terminal_log("Application closed.")
        root.destroy()


# ============================================================
# CLEAR CONTENT
# ============================================================

def clear_content():

    for widget in content.winfo_children():
        widget.destroy()


# ============================================================
# INPUT VALIDATION
# ============================================================

def numbers_only(value):

    if value == "":
        return True

    if value.isdigit():
        return True

    root.bell()

    return False


def name_only(value):

    if value == "":
        return True

    if all(
        character.isalpha() or character.isspace()
        for character in value
    ):
        return True

    root.bell()

    return False


def decimal_only(value):

    if value == "":
        return True

    # Only numbers and one decimal point
    if value.count(".") > 1:
        root.bell()
        return False

    if value == ".":
        return True

    if value.replace(".", "").isdigit():
        return True

    root.bell()

    return False


# ============================================================
# RANGE VALIDATION
# ============================================================

def check_range(entry, field_name, minimum, maximum):

    value = entry.get().strip()

    if value == "":
        return

    try:
        number = float(value)
    except ValueError:
        return

    if number < minimum or number > maximum:

        terminal_log(
            f"Invalid {field_name}: {number}. "
            f"Allowed range: {minimum}-{maximum}"
        )

        messagebox.showerror(
            "Invalid " + field_name,
            f"{field_name} must be between {minimum} and {maximum}."
        )

        entry.delete(0, tk.END)
        entry.focus()


# ============================================================
# IMMEDIATE RANGE CHECK
# ============================================================

def immediate_range_check(
    entry,
    field_name,
    minimum,
    maximum
):

    value = entry.get().strip()

    if value == "":
        return

    try:
        number = float(value)
    except ValueError:
        return

    if number > maximum:

        terminal_log(
            f"Invalid {field_name} entered: {value}"
        )

        messagebox.showerror(
            "Invalid " + field_name,
            f"{field_name} cannot be greater than {maximum}."
        )

        entry.delete(0, tk.END)
        entry.focus()


# ============================================================
# EXCEL
# ============================================================

def create_excel_file():

    if not os.path.exists(FILE_NAME):

        terminal_log("Creating Excel database.")

        workbook = openpyxl.Workbook()

        sheet = workbook.active

        sheet.title = "Student Details"

        sheet.append([
            "Student ID",
            "Student Name",
            "Attendance",
            "Study Hours",
            "Internal Marks",
            "Assignment Marks",
            "Previous Semester GPA",
            "Backlogs"
        ])

        workbook.save(FILE_NAME)

        workbook.close()

        terminal_log("Excel database created.")


# ============================================================
# LOAD STUDENTS
# ============================================================

def load_students():

    create_excel_file()

    workbook = openpyxl.load_workbook(
        FILE_NAME,
        data_only=True
    )

    sheet = workbook.active

    students = []

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row[0]:
            continue

        try:

            student = {

                "id": str(row[0]),

                "name": str(row[1]),

                "attendance": float(row[2]),

                "study_hours": float(row[3]),

                "internal": float(row[4]),

                "assignment": float(row[5]),

                "gpa": float(row[6]),

                "backlogs": int(row[7])

            }

            students.append(student)

        except (TypeError, ValueError, IndexError):

            continue

    workbook.close()

    # Student ID increasing order
    students.sort(
        key=lambda x: (
            0,
            int(x["id"])
        )
        if x["id"].isdigit()
        else (
            1,
            x["id"]
        )
    )

    terminal_log(
        f"{len(students)} student(s) loaded from Excel."
    )

    return students


# ============================================================
# SCORE CALCULATION
# ============================================================

def normalize_values(
    study_hours,
    previous_semester_gpa,
    backlogs
):

    study_hours_score = min(
        (study_hours / 10) * 100,
        100
    )

    gpa_score = min(
        (previous_semester_gpa / 10) * 100,
        100
    )

    backlog_score = max(
        0,
        100 - (backlogs * 20)
    )

    return (
        study_hours_score,
        gpa_score,
        backlog_score
    )


def calculate_score(student):

    (
        study_hours_score,
        gpa_score,
        backlog_score
    ) = normalize_values(
        student["study_hours"],
        student["gpa"],
        student["backlogs"]
    )

    score = (

        student["attendance"] * 0.20

        + study_hours_score * 0.15

        + student["internal"] * 0.20

        + student["assignment"] * 0.15

        + gpa_score * 0.20

        + backlog_score * 0.10

    )

    return round(score, 2)


# ============================================================
# RISK
# ============================================================

def get_risk_level(score):

    if score >= 75:

        return (
            "LOW DROPOUT RISK",
            GREEN,
            GREEN_LIGHT
        )

    elif score >= 50:

        return (
            "MEDIUM DROPOUT RISK",
            ORANGE,
            ORANGE_LIGHT
        )

    else:

        return (
            "HIGH DROPOUT RISK",
            RED,
            RED_LIGHT
        )


# ============================================================
# RECOMMENDATIONS
# ============================================================

def get_recommendations(student):

    recommendations = []

    attendance = student["attendance"]
    study = student["study_hours"]
    internal = student["internal"]
    assignment = student["assignment"]
    gpa = student["gpa"]
    backlogs = student["backlogs"]

    # Attendance

    if attendance < 50:

        recommendations.append(
            "Attendance is very low. Attend classes regularly and improve attendance immediately."
        )

    elif attendance < 75:

        recommendations.append(
            "Improve attendance to at least 75% to maintain consistent academic participation."
        )

    elif attendance < 85:

        recommendations.append(
            "Try to maintain attendance above 85% for better academic consistency."
        )


    # Study Hours

    if study < 2:

        recommendations.append(
            "Increase daily study time. Start with at least 2–3 focused hours every day."
        )

    elif study < 4:

        recommendations.append(
            "Increase daily study hours gradually and follow a consistent study schedule."
        )


    # Internal Marks

    if internal < 40:

        recommendations.append(
            "Internal marks are low. Revise important topics regularly and practice previous questions."
        )

    elif internal < 50:

        recommendations.append(
            "Improve internal examination marks through regular revision and practice."
        )

    elif internal < 70:

        recommendations.append(
            "Try to improve internal marks further by focusing on weak subjects."
        )


    # Assignment

    if assignment < 40:

        recommendations.append(
            "Assignment marks are very low. Complete all assignments on time and improve submission quality."
        )

    elif assignment < 50:

        recommendations.append(
            "Complete assignments regularly and improve assignment marks."
        )

    elif assignment < 70:

        recommendations.append(
            "Maintain consistency in assignments and try to score higher."
        )


    # GPA

    if gpa < 5:

        recommendations.append(
            "Previous GPA is low. Focus strongly on core subjects and create a weekly improvement plan."
        )

    elif gpa < 6:

        recommendations.append(
            "Focus on improving previous semester GPA through regular revision and subject-wise preparation."
        )

    elif gpa < 7:

        recommendations.append(
            "Maintain your academic progress and try to improve GPA further."
        )


    # Backlogs

    if backlogs >= 3:

        recommendations.append(
            "Clear pending backlogs as early as possible and give priority to failed subjects."
        )

    elif backlogs > 0:

        recommendations.append(
            "Clear pending backlogs as early as possible."
        )


    # Excellent performance

    if not recommendations:

        recommendations.append(
            "Excellent performance! Continue maintaining your attendance, study routine and academic performance."
        )

        recommendations.append(
            "Keep your assignments, internal marks and GPA at the current level or improve them further."
        )


    return recommendations


# ============================================================
# CLEAR FORM
# ============================================================

def clear_student_form():

    fields = [

        student_id_entry,
        student_name_entry,
        attendance_entry,
        study_hours_entry,
        internal_entry,
        assignment_entry,
        gpa_entry,
        backlog_entry

    ]

    for field in fields:

        if field is not None:
            field.delete(0, tk.END)

    terminal_log("Student input form cleared.")


# ============================================================
# VALIDATE STUDENT DATA
# ============================================================

def validate_student_data():

    student_id = student_id_entry.get().strip()

    student_name = student_name_entry.get().strip()

    attendance_text = attendance_entry.get().strip()

    study_hours_text = study_hours_entry.get().strip()

    internal_text = internal_entry.get().strip()

    assignment_text = assignment_entry.get().strip()

    gpa_text = gpa_entry.get().strip()

    backlog_text = backlog_entry.get().strip()


    # Student ID

    if student_id == "":

        messagebox.showerror(
            "Invalid Student ID",
            "Please enter Student ID."
        )

        student_id_entry.focus()

        return None


    # Name

    if student_name == "":

        messagebox.showerror(
            "Invalid Student Name",
            "Please enter Student Name."
        )

        student_name_entry.focus()

        return None


    # Attendance

    if attendance_text == "":

        messagebox.showerror(
            "Invalid Attendance",
            "Please enter Attendance."
        )

        attendance_entry.focus()

        return None


    attendance = float(attendance_text)

    if not 0 <= attendance <= 100:

        messagebox.showerror(
            "Invalid Attendance",
            "Attendance must be between 0 and 100."
        )

        attendance_entry.focus()

        return None


    # Study Hours

    if study_hours_text == "":

        messagebox.showerror(
            "Invalid Study Hours",
            "Please enter Study Hours."
        )

        study_hours_entry.focus()

        return None


    study_hours = float(study_hours_text)

    if not 0 <= study_hours <= 10:

        messagebox.showerror(
            "Invalid Study Hours",
            "Study Hours must be between 0 and 10."
        )

        study_hours_entry.focus()

        return None


    # Internal

    if internal_text == "":

        messagebox.showerror(
            "Invalid Internal Marks",
            "Please enter Internal Marks."
        )

        internal_entry.focus()

        return None


    internal = float(internal_text)

    if not 0 <= internal <= 100:

        messagebox.showerror(
            "Invalid Internal Marks",
            "Internal Marks must be between 0 and 100."
        )

        internal_entry.focus()

        return None


    # Assignment

    if assignment_text == "":

        messagebox.showerror(
            "Invalid Assignment Marks",
            "Please enter Assignment Marks."
        )

        assignment_entry.focus()

        return None


    assignment = float(assignment_text)

    if not 0 <= assignment <= 100:

        messagebox.showerror(
            "Invalid Assignment Marks",
            "Assignment Marks must be between 0 and 100."
        )

        assignment_entry.focus()

        return None


    # GPA

    if gpa_text == "":

        messagebox.showerror(
            "Invalid GPA",
            "Please enter GPA."
        )

        gpa_entry.focus()

        return None


    gpa = float(gpa_text)

    if not 0 <= gpa <= 10:

        messagebox.showerror(
            "Invalid GPA",
            "GPA must be between 0 and 10."
        )

        gpa_entry.focus()

        return None


    # Backlogs

    if backlog_text == "":

        messagebox.showerror(
            "Invalid Backlogs",
            "Please enter Number of Backlogs."
        )

        backlog_entry.focus()

        return None


    backlogs = int(backlog_text)


    return {

        "id": student_id,

        "name": student_name,

        "attendance": attendance,

        "study_hours": study_hours,

        "internal": internal,

        "assignment": assignment,

        "gpa": gpa,

        "backlogs": backlogs

    }


# ============================================================
# SAVE STUDENT
# ============================================================

def save_student():

    terminal_log("Save / Update button clicked.")

    student = validate_student_data()

    if student is None:
        terminal_log("Student validation failed.")
        return


    create_excel_file()

    workbook = openpyxl.load_workbook(FILE_NAME)

    sheet = workbook.active

    existing_row = None


    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        if row[0] is None:
            continue

        if str(row[0]) == student["id"]:

            existing_row = row_number

            break


    values = [

        student["id"],
        student["name"],
        student["attendance"],
        student["study_hours"],
        student["internal"],
        student["assignment"],
        student["gpa"],
        student["backlogs"]

    ]


    if existing_row is not None:

        workbook.close()

        replace = messagebox.askyesno(
            "Student ID Already Exists",
            f"Student ID {student['id']} already exists.\n\n"
            "Do you want to replace the existing student data?"
        )

        if not replace:

            terminal_log("Student update cancelled.")

            return


        workbook = openpyxl.load_workbook(FILE_NAME)

        sheet = workbook.active


        for column, value in enumerate(
            values,
            start=1
        ):

            sheet.cell(
                existing_row,
                column
            ).value = value


        workbook.save(FILE_NAME)

        workbook.close()


        terminal_log(
            f"Student ID {student['id']} updated successfully."
        )


        messagebox.showinfo(
            "Student Updated",
            f"Student ID {student['id']} updated successfully."
        )


    else:

        sheet.append(values)

        workbook.save(FILE_NAME)

        workbook.close()


        terminal_log(
            f"Student {student['id']} saved successfully."
        )


        messagebox.showinfo(
            "Student Saved",
            f"Student '{student['name']}' saved successfully."
        )


    clear_student_form()

    show_student_list()


# ============================================================
# PAGE EXIT BUTTON
# ============================================================

def create_page_exit_button(parent):

    button = tk.Button(
        parent,
        text="✕  EXIT",
        font=("Arial", 10, "bold"),
        bg=RED,
        fg=WHITE,
        activebackground="#B91C1C",
        activeforeground=WHITE,
        bd=0,
        padx=18,
        pady=8,
        cursor="hand2",
        command=exit_application
    )

    button.pack(
        side="right"
    )

    return button


# ============================================================
# PAGE HEADER
# ============================================================

def create_page_header(
    title_text,
    subtitle_text
):

    header = tk.Frame(
        content,
        bg=BG
    )

    header.pack(
        fill="x",
        pady=(0, 18)
    )


    left = tk.Frame(
        header,
        bg=BG
    )

    left.pack(
        side="left"
    )


    tk.Label(
        left,
        text=title_text,
        font=("Arial", 27, "bold"),
        bg=BG,
        fg=TEXT
    ).pack(
        anchor="w"
    )


    tk.Label(
        left,
        text=subtitle_text,
        font=("Arial", 11),
        bg=BG,
        fg=GRAY
    ).pack(
        anchor="w",
        pady=(5, 0)
    )


    create_page_exit_button(header)


# ============================================================
# CREATE FIELD
# ============================================================

def create_field(
    parent,
    row,
    column,
    label_text,
    validation=None,
    range_values=None
):

    frame = tk.Frame(
        parent,
        bg=WHITE
    )

    frame.grid(
        row=row,
        column=column,
        padx=20,
        pady=10,
        sticky="ew"
    )


    tk.Label(
        frame,
        text=label_text,
        font=("Arial", 10, "bold"),
        bg=WHITE,
        fg=GRAY
    ).pack(
        anchor="w"
    )


    options = {

        "font": ("Arial", 12),

        "bg": "#F8FAFC",

        "fg": TEXT,

        "bd": 0,

        "highlightbackground": LIGHT_GRAY,

        "highlightthickness": 1

    }


    if validation is not None:

        options["validate"] = "key"

        options["validatecommand"] = validation


    entry = tk.Entry(
        frame,
        **options
    )


    entry.pack(
        fill="x",
        ipady=8,
        pady=(5, 0)
    )


    # Immediate maximum check
    if range_values:

        minimum, maximum = range_values

        entry.bind(
            "<KeyRelease>",
            lambda event: immediate_range_check(
                entry,
                label_text.split("(")[0].strip(),
                minimum,
                maximum
            )
        )


        entry.bind(
            "<FocusOut>",
            lambda event: check_range(
                entry,
                label_text.split("(")[0].strip(),
                minimum,
                maximum
            )
        )


    return entry


# ============================================================
# STAT CARD
# ============================================================

def create_stat_card(
    parent,
    icon,
    title,
    value,
    color,
    command=None
):

    card = tk.Frame(
        parent,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1,
        cursor="hand2" if command else ""
    )


    card.pack(
        side="left",
        fill="both",
        expand=True,
        padx=6,
        ipady=18
    )


    tk.Label(
        card,
        text=icon,
        font=("Arial", 24),
        bg=WHITE
    ).pack(
        anchor="w",
        padx=20
    )


    tk.Label(
        card,
        text=title,
        font=("Arial", 9, "bold"),
        bg=WHITE,
        fg=GRAY
    ).pack(
        anchor="w",
        padx=20
    )


    tk.Label(
        card,
        text=str(value),
        font=("Arial", 25, "bold"),
        bg=WHITE,
        fg=color
    ).pack(
        anchor="w",
        padx=20
    )


    if command:

        for widget in card.winfo_children():

            widget.bind(
                "<Button-1>",
                lambda event: command()
            )

        card.bind(
            "<Button-1>",
            lambda event: command()
        )


# ============================================================
# RISK STUDENTS PAGE
# ============================================================

def show_risk_students(risk_type):

    terminal_log(
        f"Opening {risk_type} students."
    )

    clear_content()


    students = load_students()

    selected_students = []


    for student in students:

        score = calculate_score(student)

        risk, color, light = get_risk_level(score)

        if risk == risk_type:

            selected_students.append(
                (student, score, color, light)
            )


    if risk_type == "HIGH DROPOUT RISK":

        title_color = RED
        title = "High Risk Students"

    elif risk_type == "MEDIUM DROPOUT RISK":

        title_color = ORANGE
        title = "Medium Risk Students"

    else:

        title_color = GREEN
        title = "Low Risk Students"


    header = tk.Frame(
        content,
        bg=BG
    )

    header.pack(
        fill="x",
        pady=(0, 18)
    )


    left = tk.Frame(
        header,
        bg=BG
    )

    left.pack(
        side="left"
    )


    tk.Label(
        left,
        text=title,
        font=("Arial", 27, "bold"),
        bg=BG,
        fg=title_color
    ).pack(
        anchor="w"
    )


    tk.Label(
        left,
        text=f"{len(selected_students)} student(s) in {risk_type.lower()}",
        font=("Arial", 11),
        bg=BG,
        fg=GRAY
    ).pack(
        anchor="w",
        pady=(5, 0)
    )


    create_page_exit_button(header)


    # Scroll area

    outer = tk.Frame(
        content,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )

    outer.pack(
        fill="both",
        expand=True
    )


    canvas = tk.Canvas(
        outer,
        bg=WHITE,
        highlightthickness=0
    )


    scrollbar = ttk.Scrollbar(
        outer,
        orient="vertical",
        command=canvas.yview
    )


    scroll_frame = tk.Frame(
        canvas,
        bg=WHITE
    )


    scroll_frame.bind(
        "<Configure>",
        lambda event: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )


    # Create the scroll frame inside the canvas.
    # Keep the inner frame the same width as the visible canvas
    # so the student cards use the full available page width.
    scroll_window = canvas.create_window(
        (0, 0),
        window=scroll_frame,
        anchor="nw"
    )


    def resize_scroll_frame(event):

        # Expand the inner frame to the canvas width.
        # This prevents the risk-student cards from appearing short/narrow.
        canvas.itemconfigure(
            scroll_window,
            width=event.width
        )


    canvas.bind(
        "<Configure>",
        resize_scroll_frame
    )


    canvas.configure(
        yscrollcommand=scrollbar.set
    )


    canvas.pack(
        side="left",
        fill="both",
        expand=True
    )


    scrollbar.pack(
        side="right",
        fill="y"
    )


    if not selected_students:

        tk.Label(
            scroll_frame,
            text="No students found.",
            font=("Arial", 16, "bold"),
            bg=WHITE,
            fg=GRAY
        ).pack(
            pady=50
        )

        return


    for student, score, color, light in selected_students:

        card = tk.Frame(
            scroll_frame,
            bg=light,
            highlightbackground=LIGHT_GRAY,
            highlightthickness=1
        )


        card.pack(
            fill="x",
            padx=20,
            pady=10
        )


        header_card = tk.Frame(
            card,
            bg=light
        )


        header_card.pack(
            fill="x",
            padx=20,
            pady=(15, 8)
        )


        tk.Label(
            header_card,
            text=f'#{student["id"]}  {student["name"]}',
            font=("Arial", 14, "bold"),
            bg=light,
            fg=TEXT
        ).pack(
            side="left"
        )


        tk.Label(
            header_card,
            text=f"{score}/100",
            font=("Arial", 15, "bold"),
            bg=light,
            fg=color
        ).pack(
            side="right"
        )


        data = tk.Frame(
            card,
            bg=WHITE
        )


        data.pack(
            fill="x",
            padx=20,
            pady=(0, 15)
        )


        values = [

            ("Attendance", f'{student["attendance"]}%'),

            ("Study Hours", f'{student["study_hours"]} hrs/day'),

            ("Internal", f'{student["internal"]}'),

            ("Assignment", f'{student["assignment"]}'),

            ("GPA", f'{student["gpa"]}'),

            ("Backlogs", f'{student["backlogs"]}')

        ]


        for index, (label, value) in enumerate(values):

            data.columnconfigure(
                index,
                weight=1
            )


            box = tk.Frame(
                data,
                bg=WHITE
            )


            box.grid(
                row=0,
                column=index,
                padx=8,
                pady=10,
                sticky="ew"
            )


            tk.Label(
                box,
                text=label,
                font=("Arial", 8, "bold"),
                bg=WHITE,
                fg=GRAY
            ).pack()


            tk.Label(
                box,
                text=value,
                font=("Arial", 11, "bold"),
                bg=WHITE,
                fg=TEXT
            ).pack(
                pady=(3, 0)
            )


# ============================================================
# DASHBOARD
# ============================================================

def show_dashboard():

    terminal_log("Dashboard opened.")

    clear_content()


    students = load_students()

    total_students = len(students)

    high_risk = 0
    medium_risk = 0
    low_risk = 0


    for student in students:

        score = calculate_score(student)

        if score < 50:

            high_risk += 1

        elif score < 75:

            medium_risk += 1

        else:

            low_risk += 1


    create_page_header(
        "Dashboard",
        "Smart Student Dropout Risk Analytics"
    )


    cards = tk.Frame(
        content,
        bg=BG
    )


    cards.pack(
        fill="x"
    )


    create_stat_card(
        cards,
        "👨‍🎓",
        "TOTAL STUDENTS",
        total_students,
        PRIMARY
    )


    create_stat_card(
        cards,
        "🔴",
        "HIGH RISK",
        high_risk,
        RED,
        lambda: show_risk_students(
            "HIGH DROPOUT RISK"
        )
    )


    create_stat_card(
        cards,
        "🟠",
        "MEDIUM RISK",
        medium_risk,
        ORANGE,
        lambda: show_risk_students(
            "MEDIUM DROPOUT RISK"
        )
    )


    create_stat_card(
        cards,
        "🟢",
        "LOW RISK",
        low_risk,
        GREEN,
        lambda: show_risk_students(
            "LOW DROPOUT RISK"
        )
    )


    info = tk.Frame(
        content,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    info.pack(
        fill="x",
        pady=30
    )


    tk.Label(
        info,
        text="🎓 Smart Student Analytics",
        font=("Arial", 19, "bold"),
        bg=WHITE,
        fg=TEXT
    ).pack(
        anchor="w",
        padx=30,
        pady=(25, 8)
    )


    tk.Label(
        info,
        text=(
            "Click any risk card above to view students belonging "
            "to that risk category. Student IDs are displayed in "
            "increasing order."
        ),
        font=("Arial", 11),
        bg=WHITE,
        fg=GRAY,
        wraplength=900,
        justify="left"
    ).pack(
        anchor="w",
        padx=30,
        pady=(0, 25)
    )


# ============================================================
# STUDENT DETAILS
# ============================================================

def show_student_details():

    terminal_log("Student Details page opened.")

    clear_content()


    global student_id_entry
    global student_name_entry
    global attendance_entry
    global study_hours_entry
    global internal_entry
    global assignment_entry
    global gpa_entry
    global backlog_entry


    create_page_header(
        "Student Details",
        "Enter student information and save or update the student."
    )


    card = tk.Frame(
        content,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    card.pack(
        fill="x"
    )


    tk.Label(
        card,
        text="Student Information",
        font=("Arial", 16, "bold"),
        bg=WHITE,
        fg=TEXT
    ).grid(
        row=0,
        column=0,
        columnspan=3,
        sticky="w",
        padx=25,
        pady=(25, 15)
    )


    for col in range(3):

        card.columnconfigure(
            col,
            weight=1
        )


    vcmd_numbers = (
        root.register(numbers_only),
        "%P"
    )


    vcmd_name = (
        root.register(name_only),
        "%P"
    )


    vcmd_decimal = (
        root.register(decimal_only),
        "%P"
    )


    student_id_entry = create_field(
        card,
        1,
        0,
        "Student ID",
        vcmd_numbers
    )


    student_name_entry = create_field(
        card,
        1,
        1,
        "Student Name",
        vcmd_name
    )


    attendance_entry = create_field(
        card,
        1,
        2,
        "Attendance (%)  0 - 100",
        vcmd_numbers,
        (0, 100)
    )


    study_hours_entry = create_field(
        card,
        2,
        0,
        "Study Hours / Day  0 - 10",
        vcmd_decimal,
        (0, 10)
    )


    internal_entry = create_field(
        card,
        2,
        1,
        "Internal Marks  0 - 100",
        vcmd_numbers,
        (0, 100)
    )


    assignment_entry = create_field(
        card,
        2,
        2,
        "Assignment Marks  0 - 100",
        vcmd_numbers,
        (0, 100)
    )


    gpa_entry = create_field(
        card,
        3,
        0,
        "Previous Semester GPA  0 - 10",
        vcmd_decimal,
        (0, 10)
    )


    backlog_entry = create_field(
        card,
        3,
        1,
        "Number of Backlogs",
        vcmd_numbers
    )


    button_frame = tk.Frame(
        card,
        bg=WHITE
    )


    button_frame.grid(
        row=4,
        column=0,
        columnspan=3,
        pady=25
    )


    tk.Button(
        button_frame,
        text="💾  SAVE / UPDATE STUDENT",
        font=("Arial", 11, "bold"),
        bg=GREEN,
        fg=WHITE,
        activebackground="#15803D",
        activeforeground=WHITE,
        bd=0,
        padx=25,
        pady=12,
        cursor="hand2",
        command=save_student
    ).pack(
        side="left",
        padx=8
    )


    tk.Button(
        button_frame,
        text="↻  CLEAR",
        font=("Arial", 11, "bold"),
        bg=LIGHT_GRAY,
        fg=TEXT,
        bd=0,
        padx=25,
        pady=12,
        cursor="hand2",
        command=clear_student_form
    ).pack(
        side="left",
        padx=8
    )


# ============================================================
# STUDENT LIST
# ============================================================

def show_student_list():

    terminal_log("Student List page opened.")

    clear_content()


    create_page_header(
        "List of Students",
        "Students stored in student_details.xlsx • Student ID sorted"
    )


    card = tk.Frame(
        content,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    card.pack(
        fill="both",
        expand=True
    )


    style = ttk.Style()

    style.configure(
        "Treeview",
        rowheight=38,
        font=("Arial", 10)
    )


    style.configure(
        "Treeview.Heading",
        font=("Arial", 10, "bold")
    )


    columns = (

        "id",
        "name",
        "attendance",
        "study",
        "internal",
        "assignment",
        "gpa",
        "backlogs"

    )


    tree = ttk.Treeview(
        card,
        columns=columns,
        show="headings"
    )


    headings = {

        "id": "Student ID",

        "name": "Name",

        "attendance": "Attendance",

        "study": "Study Hours",

        "internal": "Internal",

        "assignment": "Assignment",

        "gpa": "GPA",

        "backlogs": "Backlogs"

    }


    widths = {

        "id": 100,

        "name": 180,

        "attendance": 110,

        "study": 110,

        "internal": 110,

        "assignment": 110,

        "gpa": 100,

        "backlogs": 100

    }


    for col in columns:

        tree.heading(
            col,
            text=headings[col]
        )

        tree.column(
            col,
            width=widths[col],
            anchor="center"
        )


    scrollbar = ttk.Scrollbar(
        card,
        orient="vertical",
        command=tree.yview
    )


    tree.configure(
        yscrollcommand=scrollbar.set
    )


    tree.pack(
        side="left",
        fill="both",
        expand=True,
        padx=15,
        pady=15
    )


    scrollbar.pack(
        side="right",
        fill="y",
        pady=15
    )


    students = load_students()


    for student in students:

        tree.insert(
            "",
            "end",
            values=(

                student["id"],

                student["name"],

                student["attendance"],

                student["study_hours"],

                student["internal"],

                student["assignment"],

                student["gpa"],

                student["backlogs"]

            )
        )


# ============================================================
# PREDICTION PAGE
# ============================================================

def show_prediction():

    terminal_log("Risk Prediction page opened.")

    clear_content()


    global student_dropdown
    global prediction_result_frame
    global prediction_recommendation_frame


    create_page_header(
        "Risk Prediction",
        "Select a saved student and predict dropout risk."
    )


    students = load_students()


    if not students:

        card = tk.Frame(
            content,
            bg=WHITE,
            highlightbackground=LIGHT_GRAY,
            highlightthickness=1
        )


        card.pack(
            fill="x",
            pady=20
        )


        tk.Label(
            card,
            text="⚠  No Saved Students",
            font=("Arial", 20, "bold"),
            bg=WHITE,
            fg=RED
        ).pack(
            pady=(40, 10)
        )


        tk.Label(
            card,
            text="Please save a student first.",
            font=("Arial", 11),
            bg=WHITE,
            fg=GRAY
        ).pack(
            pady=(0, 40)
        )


        return


    # ========================================================
    # SELECTION
    # ========================================================

    selection = tk.Frame(
        content,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    selection.pack(
        fill="x",
        pady=(0, 18)
    )


    tk.Label(
        selection,
        text="🎓  Select Student",
        font=("Arial", 13, "bold"),
        bg=WHITE,
        fg=TEXT
    ).pack(
        side="left",
        padx=(25, 15),
        pady=18
    )


    student_options = [

        f'{student["id"]} - {student["name"]}'

        for student in students

    ]


    student_dropdown = ttk.Combobox(
        selection,
        values=student_options,
        state="readonly",
        font=("Arial", 11),
        width=40
    )


    student_dropdown.pack(
        side="left",
        pady=15
    )


    tk.Button(
        selection,
        text="🔍  PREDICT",
        font=("Arial", 10, "bold"),
        bg=PRIMARY,
        fg=WHITE,
        activebackground=PRIMARY_DARK,
        activeforeground=WHITE,
        bd=0,
        padx=25,
        pady=11,
        cursor="hand2",
        command=perform_selected_prediction
    ).pack(
        side="left",
        padx=15
    )


    student_dropdown.current(0)


    # ========================================================
    # MAIN SPLIT
    # ========================================================

    main = tk.Frame(
        content,
        bg=BG
    )


    main.pack(
        fill="both",
        expand=True
    )


    main.columnconfigure(
        0,
        weight=1
    )


    main.columnconfigure(
        1,
        weight=1
    )


    main.rowconfigure(
        0,
        weight=1
    )


    # ========================================================
    # LEFT RESULT OUTER
    # ========================================================

    prediction_result_outer = tk.Frame(
        main,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    prediction_result_outer.grid(
        row=0,
        column=0,
        sticky="nsew",
        padx=(0, 8)
    )


    tk.Label(
        prediction_result_outer,
        text="Prediction Result",
        font=("Arial", 17, "bold"),
        bg=WHITE,
        fg=TEXT
    ).pack(
        anchor="w",
        padx=20,
        pady=(15, 8)
    )


    # ========================================================
    # STUDENT PERFORMANCE SCROLL AREA
    # ========================================================

    result_canvas = tk.Canvas(
        prediction_result_outer,
        bg=WHITE,
        highlightthickness=0
    )


    result_scrollbar = ttk.Scrollbar(
        prediction_result_outer,
        orient="vertical",
        command=result_canvas.yview
    )


    prediction_result_frame = tk.Frame(
        result_canvas,
        bg=WHITE
    )


    prediction_result_frame.bind(
        "<Configure>",
        lambda event: result_canvas.configure(
            scrollregion=result_canvas.bbox("all")
        )
    )


    result_canvas_window = result_canvas.create_window(
        (0, 0),
        window=prediction_result_frame,
        anchor="nw"
    )


    def resize_result_frame(event):

        result_canvas.itemconfig(
            result_canvas_window,
            width=event.width
        )


    result_canvas.bind(
        "<Configure>",
        resize_result_frame
    )


    result_canvas.configure(
        yscrollcommand=result_scrollbar.set
    )


    result_canvas.pack(
        side="left",
        fill="both",
        expand=True,
        padx=(15, 0),
        pady=(0, 15)
    )


    result_scrollbar.pack(
        side="right",
        fill="y",
        pady=(0, 15)
    )


    tk.Label(
        prediction_result_frame,
        text="Select a student and click Predict",
        font=("Arial", 11),
        bg=WHITE,
        fg=GRAY
    ).pack(
        pady=30
    )


    # ========================================================
    # RIGHT RECOMMENDATIONS
    # ========================================================

    recommendation_outer = tk.Frame(
        main,
        bg=WHITE,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    recommendation_outer.grid(
        row=0,
        column=1,
        sticky="nsew",
        padx=(8, 0)
    )


    tk.Label(
        recommendation_outer,
        text="💡  Recommendations",
        font=("Arial", 18, "bold"),
        bg=WHITE,
        fg=TEXT
    ).pack(
        anchor="w",
        padx=25,
        pady=(20, 12)
    )


    recommendation_canvas = tk.Canvas(
        recommendation_outer,
        bg=WHITE,
        highlightthickness=0
    )


    recommendation_scrollbar = ttk.Scrollbar(
        recommendation_outer,
        orient="vertical",
        command=recommendation_canvas.yview
    )


    prediction_recommendation_frame = tk.Frame(
        recommendation_canvas,
        bg=WHITE
    )


    prediction_recommendation_frame.bind(
        "<Configure>",
        lambda event: recommendation_canvas.configure(
            scrollregion=recommendation_canvas.bbox("all")
        )
    )


    recommendation_window = recommendation_canvas.create_window(
        (0, 0),
        window=prediction_recommendation_frame,
        anchor="nw"
    )


    def resize_recommendation_frame(event):

        recommendation_canvas.itemconfig(
            recommendation_window,
            width=event.width
        )


    recommendation_canvas.bind(
        "<Configure>",
        resize_recommendation_frame
    )


    recommendation_canvas.configure(
        yscrollcommand=recommendation_scrollbar.set
    )


    recommendation_canvas.pack(
        side="left",
        fill="both",
        expand=True,
        padx=(15, 0),
        pady=(0, 15)
    )


    recommendation_scrollbar.pack(
        side="right",
        fill="y",
        pady=(0, 15)
    )


# ============================================================
# PERFORM PREDICTION
# ============================================================

def perform_selected_prediction():

    terminal_log("Prediction button clicked.")

    if student_dropdown is None:
        return


    selected = student_dropdown.get()


    if not selected:

        messagebox.showwarning(
            "Select Student",
            "Please select a student."
        )

        return


    student_id = selected.split(
        " - ",
        1
    )[0]


    students = load_students()

    selected_student = None


    for student in students:

        if student["id"] == student_id:

            selected_student = student

            break


    if selected_student is None:

        messagebox.showerror(
            "Student Not Found",
            "Student data could not be found."
        )

        return


    score = calculate_score(
        selected_student
    )


    risk, risk_color, risk_light = get_risk_level(
        score
    )


    recommendations = get_recommendations(
        selected_student
    )


    terminal_log(
        f"Prediction for Student ID {student_id}: "
        f"{risk} | Score: {score}"
    )


    # ========================================================
    # CLEAR RESULT
    # ========================================================

    for widget in prediction_result_frame.winfo_children():

        widget.destroy()


    # ========================================================
    # STUDENT NAME
    # ========================================================

    tk.Label(
        prediction_result_frame,
        text=selected_student["name"],
        font=("Arial", 23, "bold"),
        bg=WHITE,
        fg=TEXT
    ).pack(
        pady=(10, 0)
    )


    tk.Label(
        prediction_result_frame,
        text=f'Student ID : {selected_student["id"]}',
        font=("Arial", 10),
        bg=WHITE,
        fg=GRAY
    ).pack(
        pady=(3, 12)
    )


    # ========================================================
    # SCORE
    # ========================================================

    score_box = tk.Frame(
        prediction_result_frame,
        bg=BLUE_LIGHT
    )


    score_box.pack(
        fill="x",
        padx=25,
        pady=5
    )


    tk.Label(
        score_box,
        text=f"{score}",
        font=("Arial", 42, "bold"),
        bg=BLUE_LIGHT,
        fg=PRIMARY
    ).pack(
        pady=(12, 0)
    )


    tk.Label(
        score_box,
        text="Performance Score / 100",
        font=("Arial", 10, "bold"),
        bg=BLUE_LIGHT,
        fg=GRAY
    ).pack(
        pady=(0, 12)
    )


    # ========================================================
    # RISK
    # ========================================================

    tk.Label(
        prediction_result_frame,
        text=risk,
        font=("Arial", 17, "bold"),
        bg=WHITE,
        fg=risk_color
    ).pack(
        pady=12
    )


    # ========================================================
    # STUDENT PERFORMANCE
    # ========================================================

    data_card = tk.Frame(
        prediction_result_frame,
        bg="#F8FAFC",
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    data_card.pack(
        fill="x",
        padx=25,
        pady=5
    )


    tk.Label(
        data_card,
        text="Student Performance",
        font=("Arial", 12, "bold"),
        bg="#F8FAFC",
        fg=TEXT
    ).pack(
        anchor="w",
        padx=18,
        pady=(12, 8)
    )


    data_values = [

        (
            "Attendance",
            f'{selected_student["attendance"]}%'
        ),

        (
            "Study Hours",
            f'{selected_student["study_hours"]} hrs/day'
        ),

        (
            "Internal Marks",
            f'{selected_student["internal"]} / 100'
        ),

        (
            "Assignment Marks",
            f'{selected_student["assignment"]} / 100'
        ),

        (
            "Previous GPA",
            f'{selected_student["gpa"]} / 10'
        ),

        (
            "Backlogs",
            f'{selected_student["backlogs"]}'
        )

    ]


    for label, value in data_values:

        row = tk.Frame(
            data_card,
            bg="#F8FAFC"
        )


        row.pack(
            fill="x",
            padx=18,
            pady=5
        )


        tk.Label(
            row,
            text=label,
            font=("Arial", 9, "bold"),
            bg="#F8FAFC",
            fg=GRAY,
            width=20,
            anchor="w"
        ).pack(
            side="left"
        )


        tk.Label(
            row,
            text=value,
            font=("Arial", 9, "bold"),
            bg="#F8FAFC",
            fg=TEXT,
            anchor="w"
        ).pack(
            side="left"
        )


    # ========================================================
    # RIGHT RECOMMENDATIONS
    # ========================================================

    for widget in prediction_recommendation_frame.winfo_children():

        widget.destroy()


    summary = tk.Frame(
        prediction_recommendation_frame,
        bg=risk_light,
        highlightbackground=LIGHT_GRAY,
        highlightthickness=1
    )


    summary.pack(
        fill="x",
        padx=8,
        pady=(5, 12)
    )


    tk.Label(
        summary,
        text=risk,
        font=("Arial", 13, "bold"),
        bg=risk_light,
        fg=risk_color
    ).pack(
        anchor="w",
        padx=15,
        pady=(12, 2)
    )


    tk.Label(
        summary,
        text=f"Performance score: {score}/100",
        font=("Arial", 10),
        bg=risk_light,
        fg=GRAY
    ).pack(
        anchor="w",
        padx=15,
        pady=(0, 12)
    )


    # ========================================================
    # RECOMMENDATION CARDS
    # ========================================================

    for index, recommendation in enumerate(
        recommendations,
        start=1
    ):

        rec_card = tk.Frame(
            prediction_recommendation_frame,
            bg=BLUE_LIGHT,
            highlightbackground="#BFDBFE",
            highlightthickness=1
        )


        rec_card.pack(
            fill="x",
            padx=8,
            pady=6
        )


        number_box = tk.Frame(
            rec_card,
            bg=WHITE
        )


        number_box.pack(
            side="left",
            padx=12,
            pady=12
        )


        tk.Label(
            number_box,
            text=str(index),
            font=("Arial", 12, "bold"),
            bg=WHITE,
            fg=PRIMARY,
            width=3
        ).pack(
            padx=3,
            pady=5
        )


        tk.Label(
            rec_card,
            text=recommendation,
            font=("Arial", 10),
            bg=BLUE_LIGHT,
            fg=TEXT,
            justify="left",
            anchor="w",
            wraplength=390
        ).pack(
            side="left",
            fill="x",
            expand=True,
            padx=(3, 12),
            pady=15
        )


# ============================================================
# SIDEBAR BUTTON
# ============================================================

def sidebar_button(
    parent,
    text,
    command
):

    button = tk.Button(
        parent,
        text=text,
        font=("Arial", 11, "bold"),
        bg=SIDEBAR,
        fg=WHITE,
        activebackground=SIDEBAR_HOVER,
        activeforeground=WHITE,
        bd=0,
        anchor="w",
        padx=25,
        pady=14,
        cursor="hand2",
        command=command
    )


    button.pack(
        fill="x",
        padx=8,
        pady=2
    )


# ============================================================
# TOP BAR
# ============================================================

topbar = tk.Frame(
    root,
    bg=SIDEBAR,
    height=65
)


topbar.pack(
    side="top",
    fill="x"
)


topbar.pack_propagate(False)


tk.Label(
    topbar,
    text="🎓",
    font=("Arial", 23),
    bg=SIDEBAR,
    fg=WHITE
).pack(
    side="left",
    padx=(25, 12)
)


tk.Label(
    topbar,
    text="SMART STUDENT ANALYTICS",
    font=("Arial", 17, "bold"),
    bg=SIDEBAR,
    fg=WHITE
).pack(
    side="left"
)


tk.Label(
    topbar,
    text="Dropout Risk Prediction System",
    font=("Arial", 9),
    bg=SIDEBAR,
    fg="#BFDBFE"
).pack(
    side="left",
    padx=15
)


# ============================================================
# SIDEBAR
# ============================================================

sidebar = tk.Frame(
    root,
    bg=SIDEBAR,
    width=235
)


sidebar.pack(
    side="left",
    fill="y"
)


sidebar.pack_propagate(False)


tk.Label(
    sidebar,
    text="MAIN MENU",
    font=("Arial", 10, "bold"),
    bg=SIDEBAR,
    fg="#93C5FD"
).pack(
    anchor="w",
    padx=25,
    pady=(30, 15)
)


sidebar_button(
    sidebar,
    "🏠   Dashboard",
    show_dashboard
)


sidebar_button(
    sidebar,
    "👤   Student Details",
    show_student_details
)


sidebar_button(
    sidebar,
    "📋   List of Students",
    show_student_list
)


sidebar_button(
    sidebar,
    "📊   Risk Prediction",
    show_prediction
)


# ============================================================
# SIDEBAR EXIT
# ============================================================

tk.Frame(
    sidebar,
    bg=SIDEBAR
).pack(
    fill="both",
    expand=True
)


tk.Button(
    sidebar,
    text="✕   EXIT APPLICATION",
    font=("Arial", 11, "bold"),
    bg=RED,
    fg=WHITE,
    activebackground="#B91C1C",
    activeforeground=WHITE,
    bd=0,
    cursor="hand2",
    command=exit_application
).pack(
    fill="x",
    padx=15,
    pady=20,
    ipady=8
)


# ============================================================
# CONTENT
# ============================================================

content = tk.Frame(
    root,
    bg=BG
)


content.pack(
    side="left",
    fill="both",
    expand=True,
    padx=30,
    pady=20
)


# ============================================================
# START APPLICATION
# ============================================================

terminal_log("Application starting...")

create_excel_file()

show_dashboard()

terminal_log("Application ready.")

root.mainloop()
