import os
import unittest
import openpyxl
import joblib


FILE_NAME = "student_details.xlsx"
MODEL_FILE = "dropout_risk_model.pkl"


def create_excel_file():
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Student Details"

        sheet.append([
            "Student ID",
            "Student Name",
            "Attendance",
            "Study Hours",
            "Internal Marks",
            "Assignment Marks",
            "Previous Semester GPA",
            "Backlogs"
        ])

        workbook.save(FILE_NAME)
        workbook.close()


def save_student(student):
    create_excel_file()

    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active

    sheet.append([
        student["id"],
        student["name"],
        student["attendance"],
        student["study_hours"],
        student["internal"],
        student["assignment"],
        student["gpa"],
        student["backlogs"]
    ])

    workbook.save(FILE_NAME)
    workbook.close()


def load_students():
    create_excel_file()

    workbook = openpyxl.load_workbook(
        FILE_NAME,
        data_only=True
    )

    sheet = workbook.active
    students = []

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):
        if not row[0]:
            continue

        students.append({
            "id": str(row[0]),
            "name": str(row[1]),
            "attendance": float(row[2]),
            "study_hours": float(row[3]),
            "internal": float(row[4]),
            "assignment": float(row[5]),
            "gpa": float(row[6]),
            "backlogs": int(row[7])
        })

    workbook.close()

    return students


def load_ml_model():
    if not os.path.exists(MODEL_FILE):
        return None

    return joblib.load(MODEL_FILE)


def predict_student(student):
    model = load_ml_model()

    if model is None:
        return None

    features = [[
        student["attendance"],
        student["study_hours"],
        student["internal"],
        student["assignment"],
        student["gpa"],
        student["backlogs"]
    ]]

    prediction = model.predict(features)[0]

    return str(prediction).upper()


def generate_recommendation(student):
    recommendations = []

    if student["attendance"] < 75:
        recommendations.append(
            "Improve attendance."
        )

    if student["study_hours"] < 4:
        recommendations.append(
            "Increase daily study hours."
        )

    if student["internal"] < 50:
        recommendations.append(
            "Improve internal marks."
        )

    if student["assignment"] < 50:
        recommendations.append(
            "Complete assignments regularly."
        )

    if student["gpa"] < 6:
        recommendations.append(
            "Focus on improving GPA."
        )

    if student["backlogs"] > 0:
        recommendations.append(
            "Clear pending backlogs."
        )

    if not recommendations:
        recommendations.append(
            "Continue maintaining good academic performance."
        )

    return recommendations


class TestExcelMLIntegration(unittest.TestCase):

    def test_student_save_and_load(self):

        student = {
            "id": "TEST001",
            "name": "Integration Student",
            "attendance": 85,
            "study_hours": 6,
            "internal": 75,
            "assignment": 80,
            "gpa": 8,
            "backlogs": 0
        }

        save_student(student)

        students = load_students()

        found = any(
            item["id"] == "TEST001"
            for item in students
        )

        self.assertTrue(found)


    def test_model_loading(self):

        model = load_ml_model()

        self.assertIsNotNone(model)


    def test_student_to_prediction(self):

        students = load_students()

        self.assertGreater(
            len(students),
            0
        )

        student = students[0]

        prediction = predict_student(
            student
        )

        self.assertIn(
            prediction,
            ["LOW", "MEDIUM", "HIGH"]
        )


    def test_prediction_and_recommendation(self):

        students = load_students()

        self.assertGreater(
            len(students),
            0
        )

        student = students[0]

        prediction = predict_student(
            student
        )

        recommendations = generate_recommendation(
            student
        )

        self.assertIsNotNone(
            prediction
        )

        self.assertIsInstance(
            recommendations,
            list
        )

        self.assertGreater(
            len(recommendations),
            0
        )


if __name__ == "__main__":
    unittest.main(
        verbosity=2
    )